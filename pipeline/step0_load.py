"""Step 0 - read the Excel workbook into a clean JSONL of paper records.

Keeping this as its own step means every later step reads a stable, plain-text
file (easy to diff, resumable, and visible in the notebook) instead of the xlsx.

    python -m pipeline.step0_load
"""
from __future__ import annotations

import argparse

from openpyxl import load_workbook

from . import config
from .util import clean_text, write_jsonl

# Spreadsheet header -> record field. Extend here if the workbook changes.
COLUMN_MAP = {
    "record_id": "record_id",
    "abstract": "abstract",
    "authors": "authors",
    "doi": "doi",
    "year": "year",
    "short_title": "short_title",
    "title": "title",
    "link": "link",
    "include": "human_include",   # the existing manual decision (may be wrong)
    "reason": "human_reason",
}


def load_papers(xlsx=config.INPUT_XLSX, sheet=config.SHEET_NAME) -> list[dict]:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb[sheet]
    headers = [clean_text(c.value) for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    records = []
    for r in range(2, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        if all(v is None for v in row):
            continue
        rec = {"row": r}
        for header, field in COLUMN_MAP.items():
            rec[field] = clean_text(row[idx[header]]) if header in idx else ""
        # A best "title" for downstream matching.
        rec["title"] = rec.get("title") or rec.get("short_title") or ""
        records.append(rec)
    return records


def main() -> None:
    ap = argparse.ArgumentParser(description="Load xlsx -> step0_papers.jsonl")
    ap.add_argument("--xlsx", default=str(config.INPUT_XLSX))
    ap.add_argument("--sheet", default=config.SHEET_NAME)
    args = ap.parse_args()

    records = load_papers(args.xlsx, args.sheet)
    write_jsonl(config.PAPERS_JSONL, records)
    with_doi = sum(1 for r in records if r["doi"])
    print(f"Loaded {len(records)} papers -> {config.PAPERS_JSONL}")
    print(f"  with DOI: {with_doi} | without DOI: {len(records) - with_doi}")


if __name__ == "__main__":
    main()
