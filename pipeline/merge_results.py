"""Merge Step-1 resolution + all backend screenings into one review workbook.

Produces '25 papers SP_screened.xlsx' with, per paper:
  - the original human include/reason
  - resolved_url + Step-1 data-quality flags
  - each backend's include / decision / reason / needs_review
  - agreement columns: do the models agree with each other, and with the human?
  - review_priority: DISAGREE / FLAGGED / REVIEW / ok  (sort on this to find mistakes)

    python -m pipeline.merge_results --backends gpt-5.5 qwen3-8b
"""
from __future__ import annotations

import argparse

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config
from .util import index_by, read_jsonl


def _human_bool(v):
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "include"):
        return True
    if s in ("false", "0", "no", "exclude"):
        return False
    return None  # e.g. 'TRUE/FALSE', blank


def build_table(backends: list[str]) -> pd.DataFrame:
    papers = read_jsonl(config.PAPERS_JSONL)
    resolved = index_by(read_jsonl(config.RESOLVED_JSONL), "record_id")
    screens = {b: index_by(read_jsonl(config.step3_jsonl(b)), "record_id") for b in backends}

    rows = []
    for p in papers:
        rid = str(p["record_id"])
        r = resolved.get(rid, {})
        human = _human_bool(p.get("human_include"))
        row = {
            "record_id": rid,
            "year": p.get("year", ""),
            "short_title": p.get("short_title", ""),
            "doi": r.get("doi") or p.get("doi", ""),
            "resolved_url": r.get("resolved_url", ""),
            "step1_flags": "; ".join(r.get("flags", [])),
            "human_include": human,
            "human_reason": p.get("human_reason", ""),
        }
        model_incs = []
        for b in backends:
            s = screens[b].get(rid, {})
            inc = s.get("include", None)
            model_incs.append(inc)
            row[f"{b}_include"] = inc
            row[f"{b}_decision"] = s.get("decision", "")
            row[f"{b}_reason"] = s.get("reason", "")
            row[f"{b}_needs_review"] = s.get("needs_human_review", "")
            row[f"{b}_error"] = s.get("error", "")

        # agreement between models (ignoring None/unclear)
        decided = [m for m in model_incs if m is not None]
        models_agree = (len(set(decided)) <= 1) if len(decided) >= 2 else None
        row["models_agree"] = models_agree
        # agreement with the human decision
        agree_human = None
        if human is not None and decided:
            agree_human = all(m == human for m in decided)
        row["agree_with_human"] = agree_human

        # review priority (what to look at first to find the mistakes)
        if models_agree is False:
            prio = "DISAGREE-MODELS"
        elif agree_human is False:
            prio = "DISAGREE-HUMAN"
        elif r.get("flags"):
            prio = "FLAGGED-STEP1"
        elif any(screens[b].get(rid, {}).get("needs_human_review") for b in backends):
            prio = "NEEDS-REVIEW"
        else:
            prio = "ok"
        row["review_priority"] = prio
        rows.append(row)

    df = pd.DataFrame(rows)
    order = {"DISAGREE-MODELS": 0, "DISAGREE-HUMAN": 1, "FLAGGED-STEP1": 2,
             "NEEDS-REVIEW": 3, "ok": 4}
    df["_o"] = df["review_priority"].map(order).fillna(9)
    df = df.sort_values(["_o", "record_id"]).drop(columns="_o").reset_index(drop=True)
    return df


def write_xlsx(df: pd.DataFrame, path=config.OUTPUT_XLSX) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "screening"

    headers = list(df.columns)
    ws.append(headers)
    for _, r in df.iterrows():
        ws.append([_cell(r[c]) for c in headers])

    # header styling
    hf = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = hf
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {"record_id": 10, "year": 8, "short_title": 46, "doi": 24,
              "resolved_url": 40, "step1_flags": 30, "human_reason": 40,
              "review_priority": 17, "human_include": 12}
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(
            h, 40 if h.endswith("_reason") else 15)

    # colour the review_priority column so mistakes jump out
    colors = {"DISAGREE-MODELS": "F8CBAD", "DISAGREE-HUMAN": "FFE699",
              "FLAGGED-STEP1": "DDEBF7", "NEEDS-REVIEW": "E2EFDA", "ok": "FFFFFF"}
    prio_col = headers.index("review_priority") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        pc = row[prio_col - 1]
        pc.fill = PatternFill(fill_type="solid",
                              fgColor=colors.get(str(pc.value), "FFFFFF"))
        pc.font = Font(bold=True)
    ws.freeze_panes = "C2"
    wb.save(path)


def _cell(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    return v


def main() -> None:
    ap = argparse.ArgumentParser(description="Merge backends -> review workbook.")
    ap.add_argument("--backends", nargs="+", default=config.DEFAULT_BACKENDS)
    ap.add_argument("--out", default=str(config.OUTPUT_XLSX))
    args = ap.parse_args()

    # only include backends that actually produced output
    have = [b for b in args.backends if config.step3_jsonl(b).exists()]
    if not have:
        raise SystemExit("No step3_*.jsonl found. Run step3_screen for a backend first.")
    df = build_table(have)
    write_xlsx(df, args.out)

    print(f"Backends merged: {have}")
    print(df["review_priority"].value_counts().to_string())
    print(f"\nWrote {args.out}")
    print("Rows needing attention (top of the sheet):")
    cols = ["record_id", "short_title", "review_priority"] + \
           [f"{b}_include" for b in have] + ["human_include"]
    print(df[df.review_priority != "ok"][cols].to_string(index=False))


if __name__ == "__main__":
    main()
